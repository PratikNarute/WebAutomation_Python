from openpyxl.reader.excel import load_workbook


@staticmethod
def importData(sheetName, rowNo, cellNo):
    wb = load_workbook(filename="C:\\Users\\Lenovo\\PycharmProjects\\WebAutomation_Python\\TestData\\testdata.xlsx")
    sheet = wb[sheetName]
    data = sheet.cell(row=rowNo, column=cellNo).value
    return data

